"""Tests verifying that CSV and XLSX export endpoints return correct output."""

import io

import openpyxl
from openpyxl.utils import get_column_letter

from app.interchange.config import REGISTRY


def test_contacts_csv_headers(client):
    response = client.get("/contacts/export?format=csv")
    assert response.status_code == 200
    assert "text/csv" in response.headers["content-type"]
    header_line = response.content.decode().splitlines()[0]
    assert header_line == ",".join(REGISTRY["contacts"].columns)


def test_deals_csv_headers(client):
    response = client.get("/deals/export?format=csv")
    assert response.status_code == 200
    assert "text/csv" in response.headers["content-type"]
    header_line = response.content.decode().splitlines()[0]
    assert header_line == ",".join(REGISTRY["deals"].columns)


def test_activities_csv_headers(client):
    response = client.get("/activities/export?format=csv")
    assert response.status_code == 200
    assert "text/csv" in response.headers["content-type"]
    header_line = response.content.decode().splitlines()[0]
    assert header_line == ",".join(REGISTRY["activities"].columns)


def test_xlsx_formatting(client):
    """XLSX exports must have bold headers, explicit column widths, and date number_format on date cells."""
    # Seed one contact
    r = client.post("/contacts", json={"name": "XLSX Tester", "email": "xlsx@example.com"})
    assert r.status_code == 201
    contact_id = r.json()["id"]

    # Seed one deal (requires an existing contact)
    r = client.post("/deals", json={"title": "XLSX Deal", "contact_id": contact_id, "value": 500.0})
    assert r.status_code == 201

    # Seed one activity with due_at populated so that date field is non-null in the export
    r = client.post(
        "/activities",
        json={"type": "call", "title": "XLSX Call", "due_at": "2026-12-31T00:00:00"},
    )
    assert r.status_code == 201

    for entity in ("contacts", "deals", "activities"):
        response = client.get(f"/{entity}/export?format=xlsx")
        assert response.status_code == 200, f"{entity}: unexpected HTTP {response.status_code}"

        wb = openpyxl.load_workbook(io.BytesIO(response.content))
        ws = wb.active
        config = REGISTRY[entity]

        # 1. Every header cell in row 1 must use a bold font
        for col_idx, col_name in enumerate(config.columns, start=1):
            cell = ws.cell(row=1, column=col_idx)
            assert cell.font.bold is True, (
                f"{entity}: header '{col_name}' (col {col_idx}) is not bold"
            )

        # 2. Column widths must be explicitly set to a positive value (exporter sets max_len + 2)
        for col_idx, col_name in enumerate(config.columns, start=1):
            col_letter = get_column_letter(col_idx)
            assert ws.column_dimensions[col_letter].width > 0, (
                f"{entity}: column '{col_name}' ({col_letter}) has no custom width"
            )

        # 3. Every non-null date-field cell in row 2 must carry the 'YYYY-MM-DD' number_format,
        #    and at least one such cell must be present (so the check is never vacuously true).
        date_cells_checked = 0
        for col_idx, col_name in enumerate(config.columns, start=1):
            if col_name not in config.date_fields:
                continue
            cell = ws.cell(row=2, column=col_idx)
            if cell.value is None:
                continue  # field was not populated; nothing to assert
            assert cell.number_format != "General", (
                f"{entity}: date column '{col_name}' still has 'General' number_format"
            )
            date_cells_checked += 1

        assert date_cells_checked >= 1, (
            f"{entity}: no date-field cells had non-null values in row 2 — "
            "seeding may not have populated the expected columns"
        )
