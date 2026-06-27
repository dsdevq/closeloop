"""Generic XLSX export for registered CRM entities."""

import io
from datetime import date, datetime

from fastapi.responses import StreamingResponse
from openpyxl import Workbook
from openpyxl.styles import Font
from openpyxl.utils import get_column_letter

from app.interchange.config import REGISTRY


def _parse_date(value: object) -> date | datetime | object:
    """Try to parse an ISO-8601 string into a date/datetime; return the value unchanged on failure."""
    if not isinstance(value, str):
        return value
    try:
        return datetime.fromisoformat(value)
    except ValueError:
        pass
    try:
        return date.fromisoformat(value)
    except ValueError:
        return value


def export_xlsx(entity: str, rows: list[dict]) -> StreamingResponse:
    """Return a StreamingResponse containing an XLSX file for the given entity.

    Columns come from REGISTRY[entity].columns; date_fields cells are written as
    Python date/datetime objects with a 'YYYY-MM-DD' number_format so Excel
    recognises them as dates rather than plain strings. The header row uses a
    bold Font. Column widths are auto-sized from the longest cell content.
    """
    config = REGISTRY[entity]
    wb = Workbook()
    ws = wb.active

    # Header row — bold font on every header cell
    for col_idx, col_name in enumerate(config.columns, start=1):
        cell = ws.cell(row=1, column=col_idx, value=col_name)
        cell.font = Font(bold=True)

    # Data rows
    for row_idx, row in enumerate(rows, start=2):
        for col_idx, col_name in enumerate(config.columns, start=1):
            value = row.get(col_name)
            if col_name in config.date_fields and value is not None:
                parsed = _parse_date(value)
                cell = ws.cell(row=row_idx, column=col_idx, value=parsed)
                cell.number_format = "YYYY-MM-DD"
            else:
                ws.cell(row=row_idx, column=col_idx, value=value)

    # Auto-size columns: max of header length and all data cell string lengths, plus padding
    for col_idx, col_name in enumerate(config.columns, start=1):
        col_letter = get_column_letter(col_idx)
        max_len = len(col_name)
        for row_idx in range(2, len(rows) + 2):
            cell_val = ws.cell(row=row_idx, column=col_idx).value
            cell_str = str(cell_val) if cell_val is not None else ""
            if len(cell_str) > max_len:
                max_len = len(cell_str)
        ws.column_dimensions[col_letter].width = max_len + 2

    buffer = io.BytesIO()
    wb.save(buffer)
    buffer.seek(0)

    return StreamingResponse(
        iter([buffer.getvalue()]),
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f"attachment; filename={entity}.xlsx"},
    )
